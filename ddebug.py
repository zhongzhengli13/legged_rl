import openpyxl
import numpy as np
import math

def diagnose(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    
    def load(sheet):
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        data = np.array([[v or 0 for v in r] 
                         for r in ws.iter_rows(min_row=2, values_only=True)], dtype=float)
        return headers, data

    print("=" * 50)
    print("【1】脚力对称性")
    _, frc = load('foot_frc')
    L, R = frc[:,0].mean(), frc[:,1].mean()
    ratio = L / max(R, 1)
    status = "✓ 正常" if 0.85 < ratio < 1.15 else ("⚠ 轻微" if 0.7 < ratio < 1.3 else "✗ 严重跛行")
    print(f"  左脚均值: {L:.1f}N  右脚均值: {R:.1f}N  比值: {ratio:.2f}  {status}")
    both = np.mean((frc[:,0]>20) & (frc[:,1]>20)) * 100
    left_only = np.mean((frc[:,0]>20) & (frc[:,1]<20)) * 100
    right_only = np.mean((frc[:,0]<20) & (frc[:,1]>20)) * 100
    print(f"  支撑分布: 左单支撑{left_only:.1f}%  右单支撑{right_only:.1f}%  双支撑{both:.1f}%")

    print("\n【2】步态相位")
    _, phs = load('foot_phs')
    diff = ((phs[:,1]-phs[:,0] + math.pi) % (2*math.pi)) - math.pi
    anti_pct = np.mean(np.abs(np.abs(diff) - math.pi) < 0.5) * 100
    status = "✓ 正常" if anti_pct > 80 else "✗ 相位混乱"
    print(f"  相位差均值: {diff.mean():.3f}  反相比例: {anti_pct:.1f}%  {status}")

    print("\n【3】奖励异常项（均值最负的前5项）")
    headers, rew = load('reward')
    means = [(headers[i], rew[:,i].mean()) for i in range(len(headers))]
    for name, val in sorted(means, key=lambda x: x[1])[:5]:
        flag = " ← 触底！" if val < -3.5 else (" ← 注意" if val < -0.5 else "")
        print(f"  {name:18s}: {val:+.4f}{flag}")

    print("\n【4】关节对称性（左右差异最大的前3个）")
    ref = [0.0, 0.0, -0.2, 0.4, -0.2, 0.0, 0.0, -0.2, 0.4, -0.2]
    names = ['L_yaw','L_roll','L_pitch','L_knee','L_ankle',
             'R_yaw','R_roll','R_pitch','R_knee','R_ankle']
    _, pos = load('joint_pos')
    diffs = []
    for i in range(5):
        d = abs(pos[:,i].mean() - pos[:,i+5].mean())
        diffs.append((names[i].replace('L_',''), d, pos[:,i].mean(), pos[:,i+5].mean(), ref[i]))
    for name, d, lv, rv, r in sorted(diffs, key=lambda x:-x[1])[:3]:
        print(f"  {name:8s}: 左{lv:+.3f} 右{rv:+.3f} 参考{r:+.3f}  差异={d:.3f}")

    print("=" * 50)

diagnose('/home/lzz/下载/github/legged_rl/experiments/try_x7/debug/debug_0.xlsx')  # 换成你的文件路径